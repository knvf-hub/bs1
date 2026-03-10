from __future__ import annotations

from io import BytesIO

from fastapi import APIRouter, File, HTTPException, UploadFile
from openpyxl import load_workbook
from pydantic import ValidationError

from app.application.product_prompts import (
    generate_product_prompts,
    generate_product_prompts_from_rows,
)
from app.domain.product_prompts import (
    BatchGenerateProductPromptsItem,
    BatchGenerateProductPromptsResponse,
    GenerateProductPromptsRequest,
    GenerateProductPromptsResponse,
)

router = APIRouter(prefix="/product-prompts", tags=["product-prompts"])


@router.post("/generate", response_model=GenerateProductPromptsResponse)
def post_generate_product_prompts(
    req: GenerateProductPromptsRequest,
) -> GenerateProductPromptsResponse:
    return generate_product_prompts(req)


@router.post("/generate-from-xlsx", response_model=BatchGenerateProductPromptsResponse)
async def post_generate_product_prompts_from_xlsx(
    file: UploadFile = File(...),
) -> BatchGenerateProductPromptsResponse:
    if not file.filename:
        raise HTTPException(status_code=400, detail="File is required")

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    try:
        content = await file.read()
        items = _parse_xlsx_rows(content)

        if not items:
            raise HTTPException(status_code=400, detail="No valid rows found in xlsx")

        return generate_product_prompts_from_rows(items)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process xlsx: {str(e)}")


def _parse_xlsx_rows(content: bytes) -> list[BatchGenerateProductPromptsItem]:
    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    header_index = {header: idx for idx, header in enumerate(headers)}

    required_headers = ["No.", "Name", "Link", "Target", "language"]
    for header in required_headers:
        if header not in header_index:
            raise HTTPException(status_code=400, detail=f"Missing required header: {header}")

    pic_header = _find_first_header(
        header_index,
        ["pic_path", "pic", "image_path", "image", "PicPath", "Pic"],
    )

    if pic_header is None:
        raise HTTPException(
            status_code=400,
            detail="Missing required header for image path: pic_path / pic / image_path / image",
        )

    items: list[BatchGenerateProductPromptsItem] = []

    for row_number, row in enumerate(rows[1:], start=2):
        no = _cell_to_str(row, header_index["No."])
        name = _cell_to_str(row, header_index["Name"])
        link = _cell_to_str(row, header_index["Link"])
        target = (_cell_to_str(row, header_index["Target"]) or "shopee").strip().lower()
        language = (_cell_to_str(row, header_index["language"]) or "th").strip().lower()
        pic_path = _cell_to_str(row, header_index[pic_header])

        if not no and not name and not link and not pic_path:
            continue

        if not pic_path:
            continue

        if not link:
            raise HTTPException(status_code=400, detail=f"Row {row_number}: missing Link")

        try:
            item = BatchGenerateProductPromptsItem(
                no=no or None,
                name=name or None,
                link=link,
                target=target,
                language=language,
                pic_path=pic_path,
            )
        except ValidationError as e:
            raise HTTPException(
                status_code=400,
                detail=f"Row {row_number}: invalid data - {e.errors()}",
            )

        items.append(item)

    return items


def _find_first_header(header_index: dict[str, int], candidates: list[str]) -> str | None:
    lowered = {key.lower(): key for key in header_index.keys()}
    for candidate in candidates:
        found = lowered.get(candidate.lower())
        if found:
            return found
    return None


def _cell_to_str(row: tuple, index: int) -> str:
    if index >= len(row):
        return ""

    value = row[index]
    if value is None:
        return ""

    return str(value).strip()